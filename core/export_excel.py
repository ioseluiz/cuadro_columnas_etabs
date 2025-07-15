from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import math
from pathlib import Path
import os
import json
import pandas as pd
from collections import defaultdict

# --- Constantes y Estilos de Borde (sin cambios) ---
REBAR_PROPERTIES_MM = [
    {'type': '#3', 'diameter': 9.525}, {'type': '#4', 'diameter': 12.7},
    {'type': '#5', 'diameter': 15.875}, {'type': '#6', 'diameter': 19.05},
    {'type': '#7', 'diameter': 22.225}, {'type': '#8', 'diameter': 25.40},
    {'type': '#9', 'diameter': 28.65}, {'type': '#10', 'diameter': 32.26},
    {'type': '#11', 'diameter': 35.81}, {'type': '#14', 'diameter': 43.00},
]
HEADER_BORDER = Border(top=Side(style='thin'), bottom=Side(style='thin'))
TOP_BORDER = Border(top=Side(style='thin'))
BOTTOM_BORDER = Border(bottom=Side(style='thin'))
thin_side = Side(border_style="thin", color="000000")
DIAGONAL_BORDER = Border(top=Side(style='thin'), bottom=Side(style='thin'), diagonalUp=True, diagonalDown=True, diagonal=thin_side)
MAX_DISTANCIA_LIBRE_A_BARRA_APOYADA_MM = 150
MAX_DISTANCIA_LIBRE_A_BARRA_APOYADA_PULGADAS = 6.0

# --- Funciones de Ayuda (sin cambios) ---
def set_border(ws, cell_range, border_style):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border_style

def agrupar_gridlines_por_contenido(datos):
    gridlines_agrupados = defaultdict(list)
    for fila in datos:
        gridlines_agrupados[fila["GridLine"]].append(fila)
    firmas = {}
    for grid, filas in gridlines_agrupados.items():
        filas_hasheables = [tuple(sorted(fila.items())) for fila in filas]
        firma_contenido = tuple(sorted(filas_hasheables))
        firmas[grid] = firma_contenido
    grupos_por_firma = defaultdict(list)
    for grid, firma in firmas.items():
        grupos_por_firma[firma].append(grid)
    resultado_final = []
    contador_grupo = 1
    for lista_de_grids in grupos_por_firma.values():
        grupo_dict = {"grupo": f"Grupo {contador_grupo}", "gridlines_iguales": sorted(lista_de_grids)}
        resultado_final.append(grupo_dict)
        contador_grupo += 1
    return resultado_final

def get_diameter(rebar):
    for bar in REBAR_PROPERTIES_MM:
        if bar['type'] == rebar:
            return bar['diameter']
    return 0

def calcular_max_espaciamiento_apoyo_lateral_cols(diametro_barra_longitudinal, unidades="mm"):
    if unidades.lower() not in ["mm", "pulgadas"]: return None
    d_b = diametro_barra_longitudinal
    dos_veces_max_dist_libre = 2 * MAX_DISTANCIA_LIBRE_A_BARRA_APOYADA_MM if unidades.lower() == "mm" else 2 * MAX_DISTANCIA_LIBRE_A_BARRA_APOYADA_PULGADAS
    hx_max = 2 * d_b + dos_veces_max_dist_libre
    return hx_max

def calcular_numero_patas_estribo_por_direccion(dimension_columna_cara, recubrimiento_a_estribo, diametro_estribo, hx_max_apoyo_lateral, unidades_dimensiones="mm"):
    if hx_max_apoyo_lateral <= 0: return None
    distancia_efectiva_apoyo = dimension_columna_cara - 2 * recubrimiento_a_estribo - diametro_estribo
    if distancia_efectiva_apoyo <= 0: return 2, 0, 0.0
    if distancia_efectiva_apoyo <= hx_max_apoyo_lateral:
        return 2, 0, distancia_efectiva_apoyo
    else:
        num_vanos = math.ceil(distancia_efectiva_apoyo / hx_max_apoyo_lateral)
        num_total_patas = int(num_vanos + 1)
        num_patas_internas_adicionales = num_total_patas - 2
        espaciamiento_real_patas = distancia_efectiva_apoyo / num_vanos
        return num_total_patas, num_patas_internas_adicionales, espaciamiento_real_patas

def calcular_espaciamiento_estribos_confinamiento_columnas_aci_318_19(menor_dimension_columna, diametro_barra_longitudinal_mas_pequena, fy_barra_longitudinal, hx, unidades="mm", fy_units="MPa"):
    if unidades.lower() not in ["mm", "pulgadas"]: return None
    if fy_units.lower() not in ["mpa", "ksi"]: return None
    fy_long_mpa = fy_barra_longitudinal * 6.89476 if fy_units.lower() == "ksi" else fy_barra_longitudinal
    if unidades.lower() == "mm":
        lower_bound_c_val, upper_bound_c_val, eq_const_c1, eq_const_c2 = 100.0, 150.0, 100.0, 350.0
    else:
        lower_bound_c_val, upper_bound_c_val, eq_const_c1, eq_const_c2 = 4.0, 6.0, 4.0, 14.0
    s_a = menor_dimension_columna / 4.0
    s_b = 5.0 * diametro_barra_longitudinal_mas_pequena if fy_long_mpa >= 550.0 else 6.0 * diametro_barra_longitudinal_mas_pequena
    val_eq_c = eq_const_c1 + (eq_const_c2 - hx) / 3.0
    s_c_temp = max(val_eq_c, lower_bound_c_val)
    s_c = min(s_c_temp, upper_bound_c_val)
    s_o_final = min(s_a, s_b, s_c)
    criterios_calculados = {"s_a (menor_dim/4)": s_a, "s_b (mult_diam_long_min)": s_b, "s_c (Eq. 18.7.5.3c ajustada)": s_c}
    return s_o_final, criterios_calculados

def calcular_lo_aci_318_19(mayor_dimension_section_col, luz_col, unidades="mm"):
    if unidades.lower() not in ["mm", "pulgadas"]: return None
    criterio_1 = mayor_dimension_section_col
    criterio_2 = luz_col / 6
    criterio_3 = 450.0 if unidades.lower() == "mm" else 18.0
    return max(criterio_1, criterio_2, criterio_3)

def get_excel_row(row_data, value):
    for item in row_data:
        if item['level'] == value: return item['row']
    return None

def get_excel_col(col_data, value):
    for item in col_data:
        if item['gridline'] == value: return item['excel_col']
    return None

def detectar_bxh_empty(work_sheet, stories_reverse, grid_lines):
    # Esta función ahora necesitará la estructura `col_rows` para saber las filas correctas
    for group in stories_reverse:
        row = group['row']
        counter_col = 0
        for col in grid_lines:
            actual_cell = work_sheet.cell(row=row, column=3 + counter_col)
            if actual_cell.value is None:
                work_sheet.merge_cells(start_row=row, start_column=3 + counter_col, end_row=row + 8, end_column=3 + counter_col)
                work_sheet.cell(row=row, column=3 + counter_col).border = DIAGONAL_BORDER
            counter_col += 1

# --- NUEVA FUNCIÓN DE AGRUPAMIENTO ---
def _agrupar_niveles_consecutivos_iguales(stories_data, column_records, grid_lines_data):
    """
    Agrupa niveles consecutivos que tienen datos de columna idénticos.
    """
    if not stories_data or not column_records:
        return []

    data_matrix = defaultdict(dict)
    for record in column_records:
        data_matrix[record['start_end_level']][record['GridLine']] = record
    
    grid_lines = [g['ID'] for g in grid_lines_data]
    memoized_signatures = {}

    def create_level_signature(level_name):
        if level_name in memoized_signatures:
            return memoized_signatures[level_name]

        signature = []
        for grid in sorted(grid_lines):
            record = data_matrix.get(level_name, {}).get(grid)
            if record:
                try:
                    rebar_diameter_mm = get_diameter(record['Rebar'])
                    depth_mm = float(record['depth']) * 10
                    width_mm = float(record['width']) * 10
                    h_floor_mm = (float(record['End Z']) - float(record['Start Z'])) * 1000

                    lo_val = calcular_lo_aci_318_19(max(depth_mm, width_mm), h_floor_mm, "mm") / 10
                    espaciamiento_val = calcular_espaciamiento_estribos_confinamiento_columnas_aci_318_19(
                        min(depth_mm, width_mm), rebar_diameter_mm, 420, 300, unidades="mm", fy_units="MPa")[0]

                    signature.append((
                        grid, record.get('bxh'), record.get('fc'), record.get('As'),
                        record.get('Rebar. Est.'), record.get('Detalle No.'),
                        round(lo_val, 2), round(espaciamiento_val, 2)
                    ))
                except (ValueError, TypeError, KeyError):
                    signature.append((grid, record.get('bxh'), None, None, None, None, None, None))
            else:
                signature.append((grid, None, None, None, None, None, None, None))
        
        result = tuple(signature)
        memoized_signatures[level_name] = result
        return result

    stories_reverse = sorted(stories_data, key=lambda x: float(x['Elevation']), reverse=True)
    levels = [f"{stories_reverse[i+1]['Name']}@{stories_reverse[i]['Name']}" for i in range(len(stories_reverse)-1)]
    
    grouped_level_info = []
    i = 0
    while i < len(levels):
        start_level_index = i
        current_level_name = levels[i]
        
        if not data_matrix.get(current_level_name):
            i += 1
            continue

        current_signature = create_level_signature(current_level_name)

        j = i + 1
        while j < len(levels):
            next_level_name = levels[j]
            if not data_matrix.get(next_level_name): break
            
            next_signature = create_level_signature(next_level_name)
            if current_signature and current_signature == next_signature:
                j += 1
            else:
                break
        
        end_level_index = j - 1
        
        start_level_name_parts = levels[start_level_index].split('@')
        end_level_name_parts = levels[end_level_index].split('@')
        
        final_level_name = f"{end_level_name_parts[0]}@{start_level_name_parts[1]}"
        
        grouped_level_info.append({
            'level': final_level_name,
            'data_source_level': current_level_name
        })
        i = j
    
    return grouped_level_info

# --- FUNCIÓN PRINCIPAL MODIFICADA ---
def generate_excel_table(folder_path, stories_data, grid_lines_data, column_records: list[dict]):
    grid_lines = [x['ID'] for x in grid_lines_data]
    wb = Workbook()
    ws = wb.active

    alinea_centrada = Alignment(horizontal='center')
    ws['A1'] = "NIVEL"; ws['A1'].border = HEADER_BORDER; ws['A1'].alignment = alinea_centrada
    ws['B1'] = "DESCRIPCION"; ws['B1'].border = HEADER_BORDER; ws['B1'].alignment = alinea_centrada
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    columns_records_reduced = [rec for rec in column_records if rec['nivel start'] != rec['nivel end']]

    # --- INICIO DE LA MODIFICACIÓN ---
    # 1. Agrupar niveles antes de generar las filas de Excel
    grouped_levels = _agrupar_niveles_consecutivos_iguales(stories_data, columns_records_reduced, grid_lines_data)

    col_rows = []
    current_excel_row = 2
    for i, group in enumerate(grouped_levels):
        level_name = group['level']
        col_rows.append({'level': level_name, 'row': current_excel_row, 'data_source_level': group['data_source_level']})

        ws.cell(row=current_excel_row, column=1).border = TOP_BORDER
        ws.cell(row=current_excel_row, column=2).border = TOP_BORDER
        ws.cell(row=current_excel_row, column=1).value = level_name
        ws.cell(row=current_excel_row, column=2).value = "b x h"
        ws.cell(row=current_excel_row + 1, column=2).value = "f'c"
        ws.cell(row=current_excel_row + 2, column=2).value = "As"
        ws.cell(row=current_excel_row + 3, column=2).value = "Est. en Lo"
        ws.cell(row=current_excel_row + 4, column=2).value = "Est. en Resto"
        ws.cell(row=current_excel_row + 5, column=2).value = "Estribo Externo"
        ws.cell(row=current_excel_row + 6, column=2).value = "Estribos Interno"
        ws.cell(row=current_excel_row + 7, column=2).value = "Lo"
        ws.cell(row=current_excel_row + 8, column=2).value = "Detalle"

        if i == len(grouped_levels) - 1:
            ws.cell(row=current_excel_row + 8, column=1).border = BOTTOM_BORDER
            ws.cell(row=current_excel_row + 8, column=2).border = BOTTOM_BORDER
            for k in range(len(grid_lines)):
                ws.cell(row=current_excel_row + 8, column=3 + k).border = BOTTOM_BORDER
        
        current_excel_row += 9
    # --- FIN DE LA MODIFICACIÓN ---

    gridline_columns = []
    counter_col = 0
    for x in grid_lines:
        ws.cell(row=1, column=3 + counter_col).value = f"{x}"
        ws.cell(row=1, column=3 + counter_col).border = TOP_BORDER
        for group in col_rows:
            ws.cell(row=group['row'], column=3 + counter_col).border = TOP_BORDER
        gridline_columns.append({'gridline': x, 'excel_col': 3 + counter_col})
        counter_col += 1
    
    # 2. Modificar el bucle de llenado de datos
    data_matrix = defaultdict(dict)
    for rec in columns_records_reduced:
        data_matrix[rec['start_end_level']][rec['GridLine']] = rec

    for group_info in col_rows:
        excel_row_start = group_info['row']
        data_source_level = group_info['data_source_level']
        
        for grid_id in grid_lines:
            record = data_matrix.get(data_source_level, {}).get(grid_id)
            excel_column = get_excel_col(gridline_columns, grid_id)
            
            if record and excel_column:
                ws.cell(row=excel_row_start, column=excel_column).value = record['bxh']
                ws.cell(row=excel_row_start + 1, column=excel_column).value = record['fc']
                ws.cell(row=excel_row_start + 2, column=excel_column).value = record['As']
                ws.cell(row=excel_row_start + 5, column=excel_column).value = record['Rebar. Est.']
                ws.cell(row=excel_row_start + 6, column=excel_column).value = record['Rebar. Est.']
                ws.cell(row=excel_row_start + 8, column=excel_column).value = record['Detalle No.']
                
                try:
                    rebar_diameter_mm = get_diameter(record['Rebar'])
                    depth_mm = float(record['depth']) * 10
                    width_mm = float(record['width']) * 10
                    h_floor_mm = (float(record['End Z']) - float(record['Start Z'])) * 1000
                    
                    lo_cm = calcular_lo_aci_318_19(max(depth_mm, width_mm), h_floor_mm, "mm") / 10
                    ws.cell(row=excel_row_start + 7, column=excel_column).value = lo_cm

                    espaciamiento_mm = calcular_espaciamiento_estribos_confinamiento_columnas_aci_318_19(
                        min(depth_mm, width_mm), rebar_diameter_mm, 420, 300, unidades="mm", fy_units="MPa")[0]
                    ws.cell(row=excel_row_start + 3, column=excel_column).value = espaciamiento_mm
                except (ValueError, TypeError, KeyError):
                    # En caso de error en los datos, dejar las celdas calculadas en blanco
                    ws.cell(row=excel_row_start + 7, column=excel_column).value = "Error"
                    ws.cell(row=excel_row_start + 3, column=excel_column).value = "Error"

    detectar_bxh_empty(ws, col_rows, grid_lines)
                
    full_filename = str(Path(folder_path) / 'cuadro_columnas.xlsx')
    try:
        wb.save(full_filename)
        print(f"ARCHIVO EXCEL CREADO EN: {full_filename}")
    except PermissionError:
        print(f"Error: Permiso denegado. Asegúrate de que el archivo '{full_filename}' no esté abierto.")