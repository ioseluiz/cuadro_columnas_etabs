import comtypes.client
import sys
import os
import pandas as pd
from openpyxl import Workbook

from utils import extractions


# from elements.story import Story

from dxf_drawer.drawing import Drawing
from dxf_drawer.detail import Detail
from dxf_drawer.column import RectangularColumn


# -- Constantes para tipos de material
# MAT_TYPE_STEEL = 1
MAT_TYPE_CONCRETE = 2
# MAT_TYPE_NODESIGN = 3
# MAT_TYPE_ALUMINIUM = 4
# MAT_TYPE_COLDFORMED = 5
# MAT_TYPE_REBAR = 6
# MAT_TYPE_TENDON = 7
# MAT_TYPE_MASONRY = 8

def obtener_sapmodel_etabs():
    sap_model = None
    ETABSObject = None
    
    try:
        try:
            print("Intentando conectar con una instancia activa de ETABS...")
            ETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
            print("Conexión exitosa con ETABSObject.")
        except (OSError, comtypes.COMError) as e:
            print("No se pudo encontrar una instancia activa de ETABS.")
            print(f"Error: {e}")
            print("Por favor, asegúrate de que ETABS esté abierto con un modelo.")
            return None
        
        # Obtener el objecto SapModel
        sap_model = ETABSObject.SapModel
        if sap_model is None:
            print("No se pudo obtener el SapModel.")
            return sap_model
        
        print("SapModel obtenido exitosamente.")
        
    except comtypes.COMError as e:
        print(f"Error de COM interactuando con ETABS: {e}")
        # Podrías querer registrar el error completo: import traceback; traceback.print_exc()
        return None
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
        # import traceback; traceback.print_exc()
        return None
    
    return sap_model



def get_column_labels(sap_model):
    if sap_model is None:
        print("Error: El objeto SapModel proporcionado no es válido.")
        return []
    
    column_labels = []
    print("\nObteniendo lista de todos los elementos frame...")
    
    try:
        num_names, names_array_tuple, ret = sap_model.FrameObj.GetNameList()
        if ret != 0:
            print(f"Error al intentar obtener la lista de frames. Código de error: {ret_val[0]}")
            return column_labels
        
        number_of_frames = num_names
        all_frame_names_tuple = names_array_tuple
        
        if number_of_frames == 0:
            print("No se encontraron objetos de tipo frame en el modelo.")
            return column_labels
        
        for frame_name in all_frame_names_tuple:
            # SapModel.FrameObj.GetTypeOAPI(frame_name) devuelve una tupla:
            # (return_code, ObjectType_String)
            # ObjectType_String puede ser "Beam", "Column", "Brace", "Null", "Link"
            type_ret = sap_model.FrameObj.GetTypeOAPI(frame_name)
            
        return None
        
        
        
        
    except AttributeError:
        print("Error: El objeto SapModel no parece tener el método 'FrameObj' o sus sub-métodos.")
        print("Asegúrate de que el modelo esté correctamente cargado e inicializado en ETABS.")
    except Exception as e:
        print(f"Ocurrió un error inesperado al obtener labels de columnas: {e}")

def get_stories_with_elevations(sap_model):
    list_stories_elevations = []
    try:
        if sap_model is None:
            print("No se pudo obtener el SapModel.")
            return list_stories_elevations
        
        num_stories, names_stories, ret = sap_model.Story.GetNameList()
        if num_stories == 0:
            print("No se encontraron stories en el modelo.")
            return list_stories_elevations
        
        print(f"Se encontraron {num_stories} niveles en el modelo.")
        
        for nombre_story in names_stories:
            elevacion = sap_model.Story.GetElevation(nombre_story)
            
            story_info = {
                "nombre": nombre_story,
                "elevacion": elevacion
            }
            list_stories_elevations.append(story_info)
            print(f" - Story: {nombre_story}, Elevacion: {elevacion}")
    
    except Exception as e:
        print(f"Ocurrio un error inesperado: {e}")
        
    return list_stories_elevations


    

def get_rectangular_concrete_sections(sap_model):
    secciones_rect_concreto = []
    prop_frame = sap_model.PropFrame
    prop_material = sap_model.PropMaterial

    if not prop_frame or not prop_material:
        print("Error: No se pudo acceder a las propiedades de secciones o materiales.")
        return secciones_rect_concreto
    
    # 1. Obtener la lista de todos los nombres de las secciones de frame ret, num_nombres,
    # nombres_secciones = prop_frame.GetNameList()
    num_nombres, nombres_secciones, ret = prop_frame.GetNameList()
    if ret != 0 or num_nombres == 0:
        print("No se encontraron secciones de marco definidas o hubo un error al obtenerlas.")
        return secciones_rect_concreto
    
    print(f"Se encontraron {num_nombres} secciones de marco. Analizando...")

    for nombre_seccion in nombres_secciones:
        try:
            # Intentar obtener las propiedades de la seccion como rectangular
            # GetRectangle(Name, FileName, MatProp, T3, T2, Color, Notes, GUID)
            file_name, mat_prop, t3, t2, color, notes, guid, ret_rect = prop_frame.GetRectangle(nombre_seccion)
            if ret_rect == 0: # Si es 0, la seccion es rectangular
                # 3. Obtener el tipo de material de la seccion
                # GetMaterial(Name, MatType, Color, Notes, GUID)
                tipo_mat_int, _, _, _, ret_mat_details = prop_material.GetMaterial(mat_prop)

                if ret_mat_details == 0:
                    # 4. Verificar si el material es concreto
                    # Alternativamente, se podria usar GetTypeOAPI(Name) que devuelve solo el tipo
                    # tipo_mat_oapi_int, ret_tipo_oapi = prop_material.GetTypeOAPI(mat_prop)
                    # if ret_tipo_oapi == 0 and tipo_mat_oapi_int == MAT_TYPE_CONCRETE

                    if tipo_mat_int == MAT_TYPE_CONCRETE:
                        seccion_info = {
                            "Nombre": nombre_seccion,
                            "Material": mat_prop,
                            "Profundidad (T3)": t3,
                            "Ancho (T2)": t2,
                        }
                        secciones_rect_concreto.append(seccion_info)
                        print(f" Seccion rectangular de concreto encontrada: {nombre_seccion} (Material: {mat_prop}, T3={t3}, T2={t2})")

        except Exception as e:
            print(f"Excepcion al procesar la seccion: '{nombre_seccion}': {e}")
            # Esto podria suceder si una seccion tiene un nombre en la lista pero
            # no se puede consultar con GetRectangle (secciones importadas extranas o nulas)
            continue
        
        if not secciones_rect_concreto:
            print("No se encontraron secciones rectangulares de concreto.")
        else:
            print(f"\nTotal de secciones rectangulares de concreto encontradas: {len(secciones_rect_concreto)}")

    return secciones_rect_concreto


def obtener_barras_refuerzo_definidas(sap_model):
    barras_refuerzo_definidas = []
    prop_rebar = sap_model.PropRebar

    if not prop_rebar:
        print("Error: No se pudo acceder a las propiedades de las barras de refuerzo.")
        return barras_refuerzo_definidas

    # 1. Obtener la lista de todos los nombres/designaciones de las barras de refuerzo
    # GetNameList() para PropRebar devuelve (ret, NumberNames, MyNameArray)
    num_nombres, nombres_barras, ret = prop_rebar.GetNameList()

    if ret != 0:
        print(f"Error al obtener la lista de nombres de barras de refuerzo. Código: {ret}")
        return barras_refuerzo_definidas

    if num_nombres == 0:
        print("No se encontraron barras de refuerzo definidas en el modelo.")
        return barras_refuerzo_definidas

    print(f"Se encontraron {num_nombres} designaciones de barras de refuerzo. Analizando...")

    for nombre_barra in nombres_barras:
        try:
            # 2. Obtener las propiedades de cada barra de refuerzo (área y diámetro)
            # GetRebarProps(Name, Area, Diameter)
            # Las unidades de Area y Diameter serán las unidades base de la base de datos del modelo.
            area_barra, diametro_barra, ret_props = prop_rebar.GetRebarProps(nombre_barra)

            if ret_props == 0:
                barra_info = {
                    "Nombre": nombre_barra,
                    "Area": area_barra,
                    "Diametro": diametro_barra,
                }
                barras_refuerzo_definidas.append(barra_info)
                print(f"  Barra de refuerzo encontrada: {nombre_barra} (Área: {area_barra:.4f}, Diámetro: {diametro_barra:.4f})")
            else:
                print(f"  Error al obtener propiedades para la barra de refuerzo '{nombre_barra}'. Código: {ret_props}")

        except Exception as e:
            print(f"Excepción al procesar la barra de refuerzo '{nombre_barra}': {e}")
            continue

    if not barras_refuerzo_definidas:
        # Este mensaje podría ser redundante si num_nombres fue 0, pero se mantiene por si hay fallos en GetRebarProps
        print("No se pudieron extraer detalles de las barras de refuerzo definidas.")
    else:
        print(f"\nTotal de designaciones de barras de refuerzo procesadas: {len(barras_refuerzo_definidas)}")

    return barras_refuerzo_definidas



def connect_to_active_etabs_instance():
    """
    Intenta conectarse a una instancia activa de ETABS y obtener la ruta del modelo abierto.

    Returns:
        tuple: (bool, str, sap_model)
               - bool: True si la conexión y la obtención de datos fueron exitosas, False en caso contrario.
               - str: Un mensaje indicando el resultado (nombre del archivo o mensaje de error).
    """
    try:
        # Intenta obtener el objeto ETABS activo
        # El nombre "CSI.ETABS.API.ETABSObject" es el ProgID común para la API de ETABS.
        # Puede variar ligeramente dependiendo de la versión de ETABS.
        etabs_object = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
        
        # Accede al SapModel, que es el objeto principal para interactuar con el modelo
        sap_model = etabs_object.SapModel
        
        # Verifica si hay un modelo abierto
        # file_path = sap_model.GetModelFilepath()
        file_name_1 = sap_model.GetModelFilename()
        
        
        if file_name_1:
            file_name = os.path.basename(file_name_1)
            print(file_name)
            return True, f"Conectado a ETABS. Modelo Abierto: {file_name}", sap_model
        else:
            return False, "ETABS está abierto, pero no hay ningún modelo cargado actualmente.", None
            
    except OSError:
        # Esto ocurre si ETABS no está abierto o el objeto COM no se puede encontrar
        return False, "No se pudo conectar a ETABS. Asegúrate de que ETABS esté abierto y ejecutándose.", None
    except AttributeError:
        # Podría ocurrir si la API cambia o si el objeto no es el esperado
        return False, "Error al interactuar con la API de ETABS. ¿Hay un modelo abierto?", None
    except Exception as e:
        # Captura cualquier otra excepción
        return False, f"Error inesperado al conectar con ETABS: {str(e)}", None


def get_excel_row(row_data, value):
    for item in row_data:
        if item['level'] == value:
            return item['row']
        
    return None

def get_excel_col(col_data, value):
    for item in col_data:
        if item['gridline'] == value:
            return  item['excel_col']
        
    return None

def create_dxf_file(column_data: list[dict]):
    pass


def generate_excel_table(stories_data, grid_lines, column_records: list[dict]):
    
    stories_reverse = []
    
    
    for story in stories_data[::-1]:
        stories_reverse.append(story)
        
    wb = Workbook()
    ws = wb.active
    
    # Create headers
    ws['A1'] = "NIVEL"
    ws['B1'] = "DESCRIPCION"
    
    
    col_rows = []
    for item in range(0,len(stories_reverse)-1):
        col_rows.append(
            {'level':f"{stories_reverse[item+1]['name']}@{stories_reverse[item]['name']}",
             'row': 9*item+2 }
        )
        ws.cell(row=9*item+2, column=1).value = f"{stories_reverse[item+1]['name']}@{stories_reverse[item]['name']}"
        ws.cell(row=9*item+2, column=2).value = "b x h"
        ws.cell(row=9*item+3, column=2).value = "f'c"
        ws.cell(row=9*item+4, column=2).value = "As"
        ws.cell(row=9*item+5, column=2).value = "Est. en Lo"
        ws.cell(row=9*item+6, column=2).value = "Est. en Resto"
        ws.cell(row=9*item+7, column=2).value = "Estribo Externo"
        ws.cell(row=9*item+8, column=2).value = "Estribos Interno"
        ws.cell(row=9*item+9, column=2).value = "Lo"
        ws.cell(row=9*item+10, column=2).value = "Detalle"
        
        
    # Columns Data
    gridline_columns = []
    counter_col = 0
    for x in grid_lines:
        ws.cell(row=1, column=3+counter_col).value = f"C-{x}"
        gridline_columns.append(
            {'gridline': x, 'excel_col': 3+counter_col}
        )
        counter_col += 1

    # Column Dataframe
    for record in column_records:
            excel_column = get_excel_col(gridline_columns,record['GridLine'])
            excel_row = get_excel_row(col_rows, record['start_end_level'])
            if excel_row:
                if excel_column:
                    #bxh
                    ws.cell(row=excel_row, column=excel_column).value = record['bxh']
                    #f'c
                    ws.cell(row=excel_row+1, column=excel_column).value = record['material']
                    # As
                    ws.cell(row=excel_row+2, column=excel_column).value = record['As']
                    # Detalle
                    ws.cell(row=excel_row+8, column=excel_column).value = record['detail']
                    




    wb.save('cuadro_columnas.xlsx')

def get_story_by_elevation(stories_data, elevation):
    for story in stories_data:
        if story['elevation'] == elevation:
            return story['name']
    return None


def get_open_model_data(SapModel):
    data_output = []
     # Get the model's name to verify the connection
    ModelName = SapModel.GetModelFilename()
    
    print(f"Model loaded: {ModelName}")

    print("\nFetching story information...")
    
    materials_dict = extractions.get_all_materials(SapModel)
    
    rebar_info = extractions.get_all_rebars(SapModel)
    print(rebar_info)
    
    
    stories = extractions.get_story_data(SapModel)
    print(stories)
    
    if not stories:
        print("Could not retrieve story information. Exiting.")
                # Optional: Release COM objects
                # sap_model = None
                # if 'etabs_object' in locals() and etabs_object is not None:
                #     etabs_object = None
        sys.exit()
        
    data_stories = []
    counter_stories = 0
        
    for story in stories:
        counter_stories += 1
        data_stories.append(Story(id=counter_stories, name=story['name'], elevation=story['elevation']))
        # print(f"  Story: {story['name']}, Elevation: {story['elevation']:.2f}")
        
    columns_at_levels = extractions.extract_columns_by_level(SapModel,stories)
    print(columns_at_levels)
    
    if columns_at_levels:
        print("\nExtracting columns for each level...")
        found_any_columns = False
        for story_info in stories:
           
            story_name = story_info["name"]
            if story_name in columns_at_levels and columns_at_levels[story_name]:
                found_any_columns = True
                print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                for col_name in sorted(columns_at_levels[story_name]):
                    info = {}
                    col_section = SapModel.FrameObj.GetSection(col_name)[0]
                    col_label, col_story, ret_label = SapModel.FrameObj.GetLabelFromName(col_name)
                    material_defined = SapModel.PropFrame.GetMaterial(col_section)[0]
                    col_point1, col_point2,  ret_points = SapModel.FrameObj.GetPoints(col_name)
                    col_x_pos = SapModel.PointObj.GetCoordCartesian(col_point1)[0]
                    col_y_pos =SapModel.PointObj.GetCoordCartesian(col_point1)[1]
                    col_z_start = SapModel.PointObj.GetCoordCartesian(col_point1)[2]
                    col_z_end = SapModel.PointObj.GetCoordCartesian(col_point2)[2]
                    col_type_enum = SapModel.PropFrame.GetTypeOAPI(col_section)[0]
                    col_shape = None
                    if col_type_enum == 8:
                        col_shape = "Rectangular"
                        width, depth = SapModel.PropFrame.GetRectangle(col_section)[2:4]
                        
                        rebar_data = extractions.get_rebar_data(SapModel, col_section, col_name)
                    elif col_type_enum == 9:
                        col_shape = "Circular"
                        width = SapModel.PropFrame.GetCircle(col_section)[2]
                        rebar_data = extractions.get_rebar_data(SapModel, col_section,col_name)
                        # print(SapModel.PropFrame.GetRebarType(col))
                        depth =None
                    elif col_type_enum == 28:
                        col_shape = "L"
                        width = None
                        depth = None
                        rebar_data = None
                    
                    #print(f"  - {col_name} - {col_section},Material: {material_defined},  Pos X: {col_x_pos}, Pos Y: {col_y_pos}, Elev: {col_z_start} to Elev: {col_z_end}, shape: {col_shape}, Width: {width}, Depth: {depth}")
                    info['col_id'] = col_name
                    info['label'] = col_label
                    info['section'] = col_section
                    info['type'] = col_shape
                    info['width'] = width
                    info['depth'] = depth
                    info['bxh'] = f"{depth}x{width}"
                    info['material'] = material_defined
                    info['pos_x'] = round(col_x_pos,2)
                    info['pos_y'] = round(col_y_pos,2)
                    info['story'] = story_name
                    info['story_elevation'] = round(story_info['elevation'],2)
                    info['story_start'] = get_story_by_elevation(stories, col_z_start)
                    info['story_end'] = get_story_by_elevation(stories, col_z_end)
                    info['z_start'] = round(col_z_start,2)
                    info['z_end'] = round(col_z_end,2)
                    info['start_end_level'] = f"{info['story_start']}@{info['story_end']}"
                    if rebar_data:
                        info['Mat. Rebar'] = rebar_data['mat_rebar_long']
                        info['cover'] = rebar_data['cover']
                        info['number_r2_bars'] = rebar_data['number_r2_bars']
                        info['number_r3_bars'] = rebar_data['number_r3_bars']
                        info['# Bars'] = rebar_data['number_bars']
                        info['Rebar'] = rebar_data['rebar_type']
                        info['Mat. Estribo'] = rebar_data['mat_rebar_confine']
                        info['As']  = f"{rebar_data['number_bars']} {rebar_data['rebar_type']}"
                        
                    else:
                        info['Mat. Rebar'] = ""
                        info['cover'] = ""
                        info['number_r2_bars'] = ""
                        info['number_r3_bars'] = ""
                        info['# Bars'] = ""
                        info['Rebar'] = ""
                        info['Mat. Estribo'] = ""
                        info['As'] = ""
                    data_output.append(info)
                    
                    
                    
                # Optionally print levels with no columns found
                # else:
                #    print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                #    print("  - No columns found with top at this level.")
            
                if not found_any_columns:
                    print("\nNo columns were found associated with any story levels based on the criteria.")
            else:
                print("No column data was extracted or an error occurred.")
                
        df_columns = pd.DataFrame(data_output)

        # Unique sections
        df_section_selected = df_columns[['bxh','As']]
        df_section_unique = df_section_selected.drop_duplicates()

        section_rows = len(df_section_unique)
        detail_values = [f"DC-{i+1}" for i in range(section_rows)]
        df_section_unique['detail'] = detail_values

        

        
        # 1. Create tuple of cols "pos_x" and "pos_y" for each row.
        # This is useful to have an object that can be used to find unique combinations.
        df_columns['temp_grid'] = df_columns[['pos_x', 'pos_y']].apply(tuple, axis=1)
        
        # 2. Get unique values of these tuples and map to an integer value
        # Use factorize(), which is efficient for this. Returns
        # - an array of integers that represents the categories.
        # - an arrray of unique categories
        # Add 1 in order that identifiers start in 1 not in 0
        df_columns['GridLine'] = pd.factorize(df_columns['temp_grid'])[0] + 1
        
        grid_lines = df_columns['GridLine'].unique()
       
        
        # 3. (Optional) Delete column temp
        df_columns = df_columns.drop(columns=['temp_grid'])

        # 4. Merge with detail for section
        df_merged = pd.merge(df_columns, df_section_unique, on=['bxh', 'As'], how='left')
        # print(df_merged)
        
        
        sections = df_merged['detail'].unique()
        # print(sections)
        number_details = len(sections)
        #5. Sort rows
        df_sorted =df_merged.sort_values(by=['GridLine', 'z_start'],ascending=True)
        # Sort dataframe by pos_x, pos_y
        # df_sorted.to_excel("column_output.xlsx")
        cols_data = df_sorted.to_dict(orient='records') # List of dictionaries
        
        # Get Rectangular Sections:
        rect_sections = get_rectangular_concrete_sections(SapModel)
        rebars_defined = obtener_barras_refuerzo_definidas(SapModel)
        extracted_sections = []
        for elemento in rect_sections:
            extracted_sections.append(elemento['Nombre'])
        rebars_in_etabs = []
        for rebar in rebars_defined:
            rebars_in_etabs.append(rebar['Nombre'])
            
        
        return {'cols_data': cols_data, 'rect_sections': extracted_sections, 'rebars_defined': rebars_in_etabs}
    
    return None
        
    
    # #Close Application
    # EtabsObject.ApplicationExit(False)

    # Convert data to list of dicts
    # col_list = df_sorted.to_dict(orient='records')
    
    # Generate Cuadro de Columnas Excel
    # generate_excel_table(stories, grid_lines ,col_list)
    
    # Generate dxf section details
    # columns = []
    # for section in sections:
    #     width = df_sorted[df_sorted['detail'] == section].iloc[0]['width'] * 1000 # Convert to mm
    #     depth = df_sorted[df_sorted['detail'] == section].iloc[0]['depth'] * 1000 # Convert to mm
    #     fc = df_sorted[df_sorted['detail'] == section].iloc[0]['material']
    #     fc = int(fc[:-3])
    #     # Convert fc to kg/cm2
    #     fc = int(fc * (12*12)*(3.28*3.28)/(100*100*2.204))
    #     fc = str(fc)
        
    #     r2_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['number_r2_bars']
    #     r3_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['number_r3_bars']
    #     rebar_type = df_sorted[df_sorted['detail'] == section].iloc[0]['Rebar']
    #     number_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['# Bars']
    #     cover = df_sorted[df_sorted['detail'] == section].iloc[0]['cover'] * 1000 #Convert to mm
    #     stirrup_type = "#4"
        
        # columns.append(
        #     {'detail': section ,'column':RectangularColumn(width=depth, height=width, fc=fc, number_of_bars=number_bars, rebar_type=rebar_type, r2_bars=r2_bars, r3_bars=r3_bars, cover = cover, stirrup_type=stirrup_type),}
        # )
        
     # 1. Create list of Detail
    # list_details = []
    # start_point = (100,100)
    # counter = 0
    # width_detail = 3000
    # height_detail = 3000
        
    # for i in range(1, len(columns)+1): 
    #     actual_col = columns[i-1]['column']
    #     origin_point = (start_point[0], start_point[1] - (height_detail*counter))
    #     detail = Detail(f"{columns[i-1]['detail']}",origin_point, width_detail, height_detail)
    #     detail.set_column(actual_col)
    #     detail.set_origin_for_col(actual_col.width, actual_col.height)
    #     list_details.append( detail)
    #     counter += 1

    # drawing = Drawing(filename='detalles_cols_etabs.dxf', list_details=list_details)
    # drawing.create_dxf()

def get_model_data(model_path):
    data_output = []
    
    # Create API helper object
    helper = comtypes.client.CreateObject('ETABSv1.Helper')
    helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)
    
    EtabsObject = helper.CreateObjectProgID("CSI.ETABS.API.ETABSObject")

    # Start ETABS application
    EtabsObject.ApplicationStart()
    # Create SnapModel Object
    SapModel =EtabsObject.SapModel
    print(SapModel)
    SapModel.File.OpenFile(model_path)
    
    # Open and save the model
    print("ETABS model opened successfully!")

    # Get the model's name to verify the connection
    ModelName = EtabsObject.SapModel.GetModelFilename()
    
    print(f"Model loaded: {ModelName}")

    print("\nFetching story information...")
    
    materials_dict = extractions.get_all_materials(SapModel)
    
    rebar_info = extractions.get_all_rebars(SapModel)
    print(rebar_info)
    
    
    stories = extractions.get_story_data(SapModel)
    print(stories)
    
    if not stories:
        print("Could not retrieve story information. Exiting.")
                # Optional: Release COM objects
                # sap_model = None
                # if 'etabs_object' in locals() and etabs_object is not None:
                #     etabs_object = None
        sys.exit()
        
    data_stories = []
    counter_stories = 0
        
    for story in stories:
        counter_stories += 1
        data_stories.append(Story(id=counter_stories, name=story['name'], elevation=story['elevation']))
        # print(f"  Story: {story['name']}, Elevation: {story['elevation']:.2f}")
        
    columns_at_levels = extractions.extract_columns_by_level(SapModel,stories)
    print(columns_at_levels)
    
    if columns_at_levels:
        print("\nExtracting columns for each level...")
        found_any_columns = False
        for story_info in stories:
           
            story_name = story_info["name"]
            if story_name in columns_at_levels and columns_at_levels[story_name]:
                found_any_columns = True
                print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                for col_name in sorted(columns_at_levels[story_name]):
                    info = {}
                    col_label, col_story, ret_label = SapModel.FrameObj.GetLabelFromName(col_name)
                    col_section = SapModel.FrameObj.GetSection(col_name)[0]
                    material_defined = SapModel.PropFrame.GetMaterial(col_section)[0]
                    col_point1, col_point2,  ret_points = SapModel.FrameObj.GetPoints(col_name)
                    col_x_pos = SapModel.PointObj.GetCoordCartesian(col_point1)[0]
                    col_y_pos =SapModel.PointObj.GetCoordCartesian(col_point1)[1]
                    col_z_start = SapModel.PointObj.GetCoordCartesian(col_point1)[2]
                    col_z_end = SapModel.PointObj.GetCoordCartesian(col_point2)[2]
                    col_type_enum = SapModel.PropFrame.GetTypeOAPI(col_section)[0]
                    col_shape = None
                    if col_type_enum == 8:
                        col_shape = "Rectangular"
                        width, depth = SapModel.PropFrame.GetRectangle(col_section)[2:4]
                        
                        rebar_data = extractions.get_rebar_data(SapModel, col_section, col_name)
                    elif col_type_enum == 9:
                        col_shape = "Circular"
                        width = SapModel.PropFrame.GetCircle(col_section)[2]
                        rebar_data = extractions.get_rebar_data(SapModel, col_section,col_name)
                        # print(SapModel.PropFrame.GetRebarType(col))
                        depth =None
                    elif col_type_enum == 28:
                        col_shape = "L"
                        width = None
                        depth = None
                        rebar_data = None
                    
                    #print(f"  - {col_name} - {col_section},Material: {material_defined},  Pos X: {col_x_pos}, Pos Y: {col_y_pos}, Elev: {col_z_start} to Elev: {col_z_end}, shape: {col_shape}, Width: {width}, Depth: {depth}")
                    info['col_id'] = col_name
                    info['label'] = col_label
                    info['section'] = col_section
                    info['type'] = col_shape
                    info['width'] = width
                    info['depth'] = depth
                    info['bxh'] = f"{depth}x{width}"
                    info['material'] = material_defined
                    info['pos_x'] = round(col_x_pos,2)
                    info['pos_y'] = round(col_y_pos,2)
                    info['story'] = story_name
                    info['story_elevation'] = round(story_info['elevation'],2)
                    info['story_start'] = get_story_by_elevation(stories, col_z_start)
                    info['story_end'] = get_story_by_elevation(stories, col_z_end)
                    info['z_start'] = round(col_z_start,2)
                    info['z_end'] = round(col_z_end,2)
                    info['start_end_level'] = f"{info['story_start']}@{info['story_end']}"
                    if rebar_data:
                        info['Mat. Rebar'] = rebar_data['mat_rebar_long']
                        info['cover'] = rebar_data['cover']
                        info['number_r2_bars'] = rebar_data['number_r2_bars']
                        info['number_r3_bars'] = rebar_data['number_r3_bars']
                        info['# Bars'] = rebar_data['number_bars']
                        info['Rebar'] = rebar_data['rebar_type']
                        info['Mat. Estribo'] = rebar_data['mat_rebar_confine']
                        info['As']  = f"{rebar_data['number_bars']} {rebar_data['rebar_type']}"
                        
                    else:
                        info['Mat. Rebar'] = ""
                        info['cover'] = ""
                        info['number_r2_bars'] = ""
                        info['number_r3_bars'] = ""
                        info['# Bars'] = ""
                        info['Rebar'] = ""
                        info['Mat. Estribo'] = ""
                        info['As'] = ""
                    data_output.append(info)
                    
                    
                    
                # Optionally print levels with no columns found
                # else:
                #    print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                #    print("  - No columns found with top at this level.")
            
                if not found_any_columns:
                    print("\nNo columns were found associated with any story levels based on the criteria.")
            else:
                print("No column data was extracted or an error occurred.")
                
        df_columns = pd.DataFrame(data_output)

        # Unique sections
        df_section_selected = df_columns[['bxh','As']]
        df_section_unique = df_section_selected.drop_duplicates()

        section_rows = len(df_section_unique)
        detail_values = [f"DC-{i+1}" for i in range(section_rows)]
        df_section_unique['detail'] = detail_values

        

        
        # 1. Create tuple of cols "pos_x" and "pos_y" for each row.
        # This is useful to have an object that can be used to find unique combinations.
        df_columns['temp_grid'] = df_columns[['pos_x', 'pos_y']].apply(tuple, axis=1)
        
        # 2. Get unique values of these tuples and map to an integer value
        # Use factorize(), which is efficient for this. Returns
        # - an array of integers that represents the categories.
        # - an arrray of unique categories
        # Add 1 in order that identifiers start in 1 not in 0
        df_columns['GridLine'] = pd.factorize(df_columns['temp_grid'])[0] + 1
        
        grid_lines = df_columns['GridLine'].unique()
       
        
        # 3. (Optional) Delete column temp
        df_columns = df_columns.drop(columns=['temp_grid'])

        # 4. Merge with detail for section
        df_merged = pd.merge(df_columns, df_section_unique, on=['bxh', 'As'], how='left')
        print(df_merged)
        
        
        sections = df_merged['detail'].unique()
        print(sections)
        number_details = len(sections)
        #5. Sort rows
        df_sorted =df_merged.sort_values(by=['GridLine', 'z_start'],ascending=True)
        # Sort dataframe by pos_x, pos_y
        df_sorted.to_excel("column_output.xlsx")
        
    
    #Close Application
    EtabsObject.ApplicationExit(False)

    # Convert data to list of dicts
    col_list = df_sorted.to_dict(orient='records')
    
    # Generate Cuadro de Columnas Excel
    generate_excel_table(stories, grid_lines ,col_list)
    
    # Generate dxf section details
    columns = []
    for section in sections:
        width = df_sorted[df_sorted['detail'] == section].iloc[0]['width'] * 1000 # Convert to mm
        depth = df_sorted[df_sorted['detail'] == section].iloc[0]['depth'] * 1000 # Convert to mm
        fc = df_sorted[df_sorted['detail'] == section].iloc[0]['material']
        fc = int(fc[:-3])
        # Convert fc to kg/cm2
        fc = int(fc * (12*12)*(3.28*3.28)/(100*100*2.204))
        fc = str(fc)
        
        r2_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['number_r2_bars']
        r3_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['number_r3_bars']
        rebar_type = df_sorted[df_sorted['detail'] == section].iloc[0]['Rebar']
        number_bars = df_sorted[df_sorted['detail'] == section].iloc[0]['# Bars']
        cover = df_sorted[df_sorted['detail'] == section].iloc[0]['cover'] * 1000 #Convert to mm
        stirrup_type = "#4"
        
        columns.append(
            {'detail': section ,'column':RectangularColumn(width=depth, height=width, fc=fc, number_of_bars=number_bars, rebar_type=rebar_type, r2_bars=r2_bars, r3_bars=r3_bars, cover = cover, stirrup_type=stirrup_type),}
        )
        
     # 1. Create list of Detail
    list_details = []
    start_point = (100,100)
    counter = 0
    width_detail = 3000
    height_detail = 3000
        
    for i in range(1, len(columns)+1): 
        actual_col = columns[i-1]['column']
        origin_point = (start_point[0], start_point[1] - (height_detail*counter))
        detail = Detail(f"{columns[i-1]['detail']}",origin_point, width_detail, height_detail)
        detail.set_column(actual_col)
        detail.set_origin_for_col(actual_col.width, actual_col.height)
        list_details.append( detail)
        counter += 1

    drawing = Drawing(filename='detalles_cols_etabs.dxf', list_details=list_details)
    drawing.create_dxf()
        